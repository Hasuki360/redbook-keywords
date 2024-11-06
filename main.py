import os
import random
import time
import threading
import sys

from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from lxml import etree

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, ElementClickInterceptedException

from LOGGER import GetLogger

logger = GetLogger(logger_name="XiaoHongShu", debug=False, log_file="XiaoHongShu.log")

# 设置 Chrome 选项以避免打印出过多日志
chrome_options = Options()
USERPATH = os.path.abspath("./userData")
service = Service('./chromedriver.exe')
chrome_options.add_argument(f"--user-data-dir={USERPATH}")
chrome_options.add_argument("--headless=new")  # 无头模式
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--window-size=1920,1080")

chrome_options.add_argument("--disable-logging")
chrome_options.add_argument("--log-level=3")
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
chrome_options.add_experimental_option("useAutomationExtension", False)
# 初始化 WebDriver
driver = webdriver.Chrome(options=chrome_options, service=service)

LOGIN_CONTAINER_CLASS = 'login-container'
LOGIN_BTN_CLASS = 'login-btn'
USER_BTN_CLASS = 'user'

XIAOHONGSHU_URL = 'https://www.xiaohongshu.com/explore'

def get_sort_type():
    """获取用户选择的排序方式"""
    while True:
        sort_input = input("\n请选择排序方式（直接回车默认为'最新'）：\n1.综合\n2.最新\n3.最热\n请输入(1/2/3): ").strip()
        
        # 默认选择"最新"
        if not sort_input:
            logger.info("使用默认排序方式：最新")
            return "最新"
            
        # 映射用户输入到具体排序方式
        sort_mapping = {
            "1": "综合",
            "2": "最新",
            "3": "最热"
        }
        
        if sort_input in sort_mapping:
            selected_sort = sort_mapping[sort_input]
            logger.info(f"选择的排序方式：{selected_sort}")
            return selected_sort
        else:
            print("无效的选择！请输入1、2、3或直接回车选择默认值(最新)")

def search_page(keyword, total, exclude_keywords, sort_type):
    excluded_count = 0  # 添加计数器
    logger.info("打开小红书页面中")
    driver.get(XIAOHONGSHU_URL)
    SEARCH_INPUT_ID = 'search-input'
    SEARCH_ICON_CLASS = 'search-icon'

    # 向搜索框发送关键字
    logger.info(f"搜索关键字[{keyword}]中")
    search_input_object = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, SEARCH_INPUT_ID))
    )
    search_input_object.send_keys(keyword)
    logger.info(f"输入关键字[{keyword}]完成")

    # 点击搜索按钮
    logger.info(f"点击搜索按钮中")
    search_input_button_object = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CLASS_NAME, SEARCH_ICON_CLASS))
    )
    search_input_button_object.click()
    logger.info(f"点击搜索按钮完成")
    time.sleep(2)

    # 点击筛选按钮
    FILTER_INCO_CLASS = 'filter'
    filter_area_object = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CLASS_NAME, FILTER_INCO_CLASS))
    )
    # 悬浮至该按钮上
    ActionChains(driver).move_to_element(filter_area_object).perform()
    time.sleep(1)  # 等待下拉菜单显示

    # 获取所有具有类名 'dropdown-items' 的元素
    dropdown_menus = driver.find_elements(By.CLASS_NAME, "dropdown-items")

    # 过滤出 display 样式为 inline-block 的元素
    visible_menus = [
        menu
        for menu in dropdown_menus
        if driver.execute_script(
            "return window.getComputedStyle(arguments[0]).display;", menu
        )
        == "inline-block"
    ]

    if len(visible_menus) == 0:
        logger.error("未找到任何可见的菜单")
        return

    # 获取第一个可见的菜单
    dropdown_menu = visible_menus[0]
    # 获取所有 <li> 元素
    li_elements = dropdown_menu.find_elements(By.TAG_NAME, "li")

    # 遍历每个 <li>，查找对应的排序选项
    for li in li_elements:
        try:
            span = li.find_element(By.TAG_NAME, "span")
            # 这里判断 span.text 是否等于用户选择的排序方式
            if span.text == sort_type:
                li.click()
                logger.info(f"成功点击 '{sort_type}' 选项")
                break
        except Exception as e:
            logger.warning(f"在<li>元素中查找<span>时出现问题: {e}")
    else:
        logger.error(f"未找到包含 '{sort_type}' 文本的 <li> 标签")

    last_index = 0  # 上一次最后爬取的数据索引
    no_change_count = 0  # 记录没有变化的次数
    no_change_limit = 3  # 设定限制次数
    while last_index <= int(total - 1):
        if last_index == 0:
            logger.info("开始爬取数据中")
        else:
            logger.info(f"该翻页了! 起始数据{last_index + 1}条: 获取页面数据中")
        # 获取内容页面
        FEEDS_CONTAINER_CLASS = "feeds-container"
        feed_container_object = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, FEEDS_CONTAINER_CLASS))
        )
        # 获取容器内的所有 <section> 标签（代表每个帖子）
        sections = feed_container_object.find_elements(By.TAG_NAME, "section")
        # 检测数据变化的循环
        previous_last_index = last_index
        # 如果已经爬取了数据则需要对页面进行滚动加载操作
        if last_index != 0:
            logger.info(f"开始滚动页面")
            # 优化滑动操作
            for _ in range(3):  # 尝试最多3次滑动
                ActionChains(driver).scroll_by_amount(0, 300).perform()
                time.sleep(1)
                feed_container_object = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, FEEDS_CONTAINER_CLASS))
                )
                sections = feed_container_object.find_elements(By.TAG_NAME, "section")
                try:
                    first_section = int(sections[0].get_attribute("data-index"))
                    last_section = int(sections[-1].get_attribute("data-index"))
                    if first_section <= last_index < last_section:
                        logger.info(f"滑动成功，当前数据范围：{first_section} - {last_section}")
                        break
                except StaleElementReferenceException:
                    logger.warning("检测到 StaleElementReferenceException，重新获取 sections")
                    continue
            else:
                logger.warning("滑动操作未能达到预期效果，继续处理现有数据")

        last_index = get_container(keyword, total, sections, last_index, exclude_keywords)

        # 检查是否数据变化
        if last_index == previous_last_index:
            no_change_count += 1
            if no_change_count >= no_change_limit:
                logger.info("数据已停止变化，结束爬取")
                break
        else:
            no_change_count = 0  # 重置计数器

    logger.info(f"总共爬取 {last_index} 个帖子，其中 {excluded_count} 个因包含排除关键词而被跳过")

def get_container(keywords, total, sections, last_index, exclude_keywords):
    file = f"{keywords}.xlsx"
    # 获取内容页面
    FEEDS_CONTAINER_CLASS = "feeds-container"
    feed_container_object = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, FEEDS_CONTAINER_CLASS))
    )
    time.sleep(2)
    logger.info(f"获取页面中卡片完成")

    # 动态获取刷新以对抗爬虫过程中页面元素更新
    for i in range(len(sections)):
        if last_index == total:
            logger.info(f"所有帖子爬取完毕")
            break
        try:
            logger.info(f"读取第{last_index}个帖子")
            # 重新定位当前 section，以防止 stale element 问题
            sections = feed_container_object.find_elements(By.TAG_NAME, "section")
            # 判断是否包含 data-width 和 data-height 属性，对抗广告
            data_index = sections[i].get_attribute("data-index")
            data_width = sections[i].get_attribute("data-width")
            data_height = sections[i].get_attribute("data-height")
            if int(data_index) < int(last_index):
                logger.info(f"本批次, 第 {i} 个帖子爬取了，跳过")
                continue
            if not data_width or not data_height:
                logger.info(f"第 {last_index} 个帖子为广告，跳过")
                last_index += 1
                tree = None
            else:
                logger.info(f"点击第{last_index}个帖子")
                try:
                    sections[i].click()
                except ElementClickInterceptedException:
                    logger.error(
                        f"第 {last_index} 个帖子点击失败,可能是不兼容的直播或广告,跳过"
                    )
                    tree = None
                    last_index += 1
                    continue
                # 抓取帖子页面内容
                AUTHOR_AREA_CLASS = "author-wrapper"
                TITLE_AREA_ID = "detail-title"
                CONTENT_AREA_ID = "detail-desc"
                BOTTOM_AREA_CLASS = "bottom-container"
                INTERATE_AREA_CLASS = "interact-container"
                logger.info("加载数据中")
                wait = WebDriverWait(driver, 10)
                # 等待页面的各个部分加载完成
                wait.until(
                    EC.presence_of_element_located((By.CLASS_NAME, AUTHOR_AREA_CLASS))
                )
                logger.info("加载作者信息完成")
                try:
                    wait.until(
                        EC.presence_of_element_located((By.ID, TITLE_AREA_ID))
                    )
                    logger.info("加载标题完成")
                    wait.until(
                        EC.presence_of_element_located((By.ID, CONTENT_AREA_ID))
                    )
                    logger.info("加载内容完成")
                    wait.until(
                        EC.presence_of_element_located((By.CLASS_NAME, BOTTOM_AREA_CLASS))
                    )
                    logger.info("加载底部信息完成")
                except TimeoutException:
                    logger.warning("出现内容为空,此消息作为警告消息提示")
                except Exception as e:
                    logger.error(f"加载内容失败，错误信息：{e}")

                wait.until(
                    EC.presence_of_element_located(
                        (By.CLASS_NAME, INTERATE_AREA_CLASS)
                    )
                )
                logger.info("加载具体浏览数据完成")
                # 截取页面数据
                logger.info("截取当前页面数据中...")
                page_content = driver.page_source
                tree = etree.HTML(page_content)

                time.sleep(2)
                # 发送 ESC 键关闭帖子
                logger.info("点击关闭按钮关闭帖子")
                CLOSS_BTN_CLASS = "close-circle"
                close_btn = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, CLOSS_BTN_CLASS))
                )
                close_btn.click()
                logger.info("已关闭帖子")

                # 更新读取进度
                logger.info(f"读取第{last_index}个帖子完成")
                last_index += 1
                # 等待关闭后重新定位页面
                logger.info("重新定位页面中")
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, FEEDS_CONTAINER_CLASS))
                )
                logger.info("重新定位页面完成")
                time.sleep(3)

            get_content(tree, file, exclude_keywords)  # 获取页面数据
        except StaleElementReferenceException:
            logger.error(
                f"StaleElementReferenceException: 第{last_index}个帖子元素失效，重新尝试获取元素"
            )
            # 重新获取页面的帖子列表并重试
            feed_container_object = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, FEEDS_CONTAINER_CLASS))
            )
            sections = feed_container_object.find_elements(By.TAG_NAME, "section")
        except IndexError:
            logger.warning(
                f"已经爬取到了所有帖子无法再爬取更多了，总数为{last_index}个"
            )
        except Exception as e:
            logger.error(f"发生未预期的错误: {e}")
            continue
    logger.info(f"爬取完成,爬取到{last_index}个帖子")
    return last_index

def get_content(tree, file, exclude_keywords):
    if not tree:
        data = {"remark": "这是一个广告"}
        write_to_excel(data, file)
        return

    data = {}
    
    # 获取用户名
    username_element = tree.xpath('//div[@class="author-wrapper"]//span[@class="username"]/text()')
    if username_element:
        data["username"] = username_element[0]

    # 获取标题
    title_element = tree.xpath('//div[@id="detail-title"]/text()')
    if title_element:
        data["title"] = title_element[0]

    # 获取内容
    content_element = tree.xpath('//div[@id="detail-desc"]//span[@class="note-text"]/span/text()')
    if content_element:
        data["content"] = content_element[0]

    # 获取标签
    tags_elements = tree.xpath('//a[@class="tag"]/text()')
    tags = [tag.strip() for tag in tags_elements]
    data["tags"] = tags

    # 检查标题、内容和标签是否包含需要排除的关键词
    text_to_check = [
        data.get("title", ""),
        data.get("content", ""),
        *tags
    ]
    
    for text in text_to_check:
        if any(keyword in text for keyword in exclude_keywords):
            logger.info(f"发现排除关键词，跳过此帖子")
            return  # 直接返回，不写入Excel

    # 获取其他数据...
    date_local_element = tree.xpath('//div[@class="bottom-container"]//span[@class="date"]/text()')
    if date_local_element:
        data["date_local"] = date_local_element[0]

    like_count_element = tree.xpath('//div[@class="interact-container"]/div/div//span[contains(@class, "like-wrapper")]//span[contains(@class, "count")]/text()')
    if like_count_element:
        data["like_count"] = like_count_element[0]
        if data["like_count"] == "点赞":
            data["like_count"] = 0

    collect_count_element = tree.xpath('//span[contains(@class, "collect-wrapper")]//span[contains(@class, "count")]/text()')
    if collect_count_element:
        data["collect_count"] = collect_count_element[0]
        if data["collect_count"] == "收藏":
            data["collect_count"] = 0

    comment_count_element = tree.xpath('//span[contains(@class, "chat-wrapper")]//span[contains(@class, "count")]/text()')
    if comment_count_element:
        data["comment_count"] = comment_count_element[0]
        if data["comment_count"] == "评论":
            data["comment_count"] = 0

    # 将数据写入 Excel
    write_to_excel(data, file)

    # 随机延迟 5-15 秒
    delay = random.randint(5, 15)
    logger.info(f"等待 {delay} 秒...")
    time.sleep(delay)


def write_to_excel(data, filename="output.xlsx"):
    try:
        # 尝试打开已有文件，否则创建新文件
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        # 写入表头
        headers = [
            "用户名",
            "标题",
            "内容",
            "标签",
            "发布时间和地点",
            "点赞数",
            "收藏数",
            "评论数",
            "备注",
        ]
        ws.append(headers)
        # 设置表头样式
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 15  # 设置列宽

    # 写入数据到表格
    ws.append(
        [
            data.get("username", ""),
            data.get("title", ""),
            data.get("content", ""),
            ", ".join(data.get("tags", [])),
            data.get("date_local", ""),
            data.get("like_count", ""),
            data.get("collect_count", ""),
            data.get("comment_count", ""),
            data.get("remark", ""),
        ]
    )

    # 保存文件
    wb.save(filename)
    logger.info(f"数据已写入 {filename}")

if __name__ == "__main__":
    while True:
        try:
            # 获取用户输入的关键词，设置默认值为"运动"
            keywords = input("请输入要爬取的关键词（直接回车默认为'运动'）：").strip()
            if not keywords:
                keywords = "运动"
            
            # 获取用户输入的爬取数量，设置默认值为220
            total_input = input("请输入要爬取的最大条数（直接回车默认为220）：").strip()
            total = 220 if not total_input else int(total_input)
            
            # 设置默认的排除关键词
            default_exclude_keywords = ["宠物", "论文", "骗子"，"虚假"，"法律常识", "软文", "代发短信"]
            exclude_input = input(
                "请输入要排除的关键词（用空格分隔，直接回车使用默认关键词）：\n"
                f"默认关键词：{' '.join(default_exclude_keywords)}\n"
            ).strip()
            
            if not exclude_input:
                exclude_keywords = default_exclude_keywords
            else:
                exclude_keywords = exclude_input.split()

            # 获取排序方式
            sort_type = get_sort_type()

            # 爬取关键词
            search_page(keywords, total, exclude_keywords, sort_type)

            # 询问用户是否继续爬取，回车为继续
            continue_crawling = input("是否继续爬取？（直接回车继续，输入任意键退出）: ")
            if continue_crawling:  # 如果输入了任何内容就退出
                break
                
        except Exception as e:
            logger.error(f"发生错误: {e}")
            choice = input("是否继续程序？（直接回车继续，输入任意键退出）: ")
            if choice:  # 如果输入了任何内容就退出
                break

    # 退出浏览器
    driver.quit()
    logger.info("爬取结束！")  
